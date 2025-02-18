import os
import openpyxl.workbook
import requests
import openpyxl


WHEATHER_BASELINE = "https://api.weatherapi.com/v1"
REALTIME_END_POINT = "/current.json"
FORECAST_END_POINT = "/forecast.json"
WEATHER_API_KEY = "3cc7be82134549e1ab210308251002"  #Temp
NEWS_BASELINE = "https://newsapi.org"
NEW_EVERYTHING_END_POINT = "/v2/everything"
NEWS_API_KEY = "fcc32b925a8d4d438477d28e0dab520a"   #Temp

class Dashboard:

    def __init__(self):
        try: 
            while True: 
                try:
                    self.zipcode_input = int(input("Enter your zip code: "))
                    self.weather = Weather()  
                    self.weather.realtime_weather(self.zipcode_input)
                    self.history = History() 
                    break
                except ValueError:  
                    print("Invalid zip code! Please try again.")
            

            while True:                               
                option_input = int(input(
                "Choose an option in the menu...\n"
                "1. Add Zip code\n"
                "2. Weather\n"
                "3. News\n"
                "4. History\n"
                "5. Exit\n"
                "Option: "))

                if(option_input == 1) :
                    self.zipcode_input = self.add_zipcode() 
                    self.weather.realtime_weather(self.zipcode_input) 
                    print(f"New zip code updated: {self.zipcode_input}")
                elif(option_input == 2) :
                    print(self.weather.weather_forecast(zipcode=self.zipcode_input))
                    self.history.saved_history(option_input,self.weather.weather_forecast(zipcode=self.zipcode_input))
                elif(option_input == 3) :
                    news = News((self.weather.state + " " + self.weather.city))
                    self.history.saved_history(option_input,news.news_data())
                elif(option_input == 4) :
                    self.history.read_history()   
                elif(option_input == 5) :
                    print("Exit. Bye-Bye!!") 
                    break          


        except KeyboardInterrupt:
            print("\nCTRL+C detected. Exiting program safely. Goodbye!")

    #NOTE:ADD Zip code
    def add_zipcode(self)-> int:
        while True:    
                zipcode = int(input("Enter your zipcode : "))
                param={
                    "key" : WEATHER_API_KEY,
                    "q" :   zipcode
                }
                response = requests.get(WHEATHER_BASELINE + REALTIME_END_POINT, params=param)
                if response.status_code == 200 :
                    self.zipcode_input =zipcode
                    print("New zip code added successfully!")
                    break    
                else:
                    print("Invalid zip code! Please try again.")
        return self.zipcode_input


#NOTE: Weather request
class Weather:

    def __init__(self):
        pass

    def realtime_weather(self, zipcode) -> str:
    
        param={
            "key" : WEATHER_API_KEY,
            "q" :   zipcode
        }

        response = requests.get(WHEATHER_BASELINE + REALTIME_END_POINT, params=param)

        if response.status_code == 200:
            data = response.json()
            self.city = data["location"]["name"]
            self.state = data["location"]["region"]
            self.temp_c = data["current"]["temp_c"]
            self.temp_f = data["current"]["temp_f"]
            formatted_realtime = (
                "\nState: " + self.state +
                "\nCity: " + self.city +
                "\nCelsius: " + str(self.temp_c) +
                "\nFahrenheit: " + str(self.temp_f) + "\n"
            )
            print(formatted_realtime)
        else:
            raise ValueError("Zipcode is incorrect!!")
    
        
    def weather_forecast(self, zipcode:int) -> str:
        
        param={
            "key" : WEATHER_API_KEY,
            "q" :   zipcode,
            "days" : 3
        }
        response = requests.get(WHEATHER_BASELINE+FORECAST_END_POINT, params=param)

        data = response.json()
        day1_date = data["forecast"]["forecastday"][1]["date"]
        day1_max_c = data["forecast"]["forecastday"][1]["day"]["maxtemp_c"]
        day1_min_c = data["forecast"]["forecastday"][1]["day"]["mintemp_c"]
        day1_max_f = data["forecast"]["forecastday"][1]["day"]["maxtemp_f"]
        day1_min_f = data["forecast"]["forecastday"][1]["day"]["mintemp_f"]

        day2_date = data["forecast"]["forecastday"][2]["date"]
        day2_max_c = data["forecast"]["forecastday"][2]["day"]["maxtemp_c"]
        day2_min_c = data["forecast"]["forecastday"][2]["day"]["mintemp_c"]
        day2_max_f = data["forecast"]["forecastday"][2]["day"]["maxtemp_f"]
        day2_min_f = data["forecast"]["forecastday"][2]["day"]["mintemp_f"]
       
        data = [
        ("State", self.state),
        ("City", self.city),
        ("Date", str(day1_date)),
        ("Max_Celsius", str(day1_max_c)),
        ("Min_Celsius", str(day1_min_c)),
        ("Max_Fahrenheit", str(day1_max_f)),
        ("Min_Fahrenheit", str(day1_min_f)),
        ("Date", str(day2_date)),
        ("Max_Celsius", str(day2_max_c)),
        ("Min_Celsius", str(day2_min_c)),
        ("Max_Fahrenheit", str(day2_max_f)),
        ("Min_Fahrenheit", str(day2_min_f))
        ]
        return data

#NOTE: NEWS request
class News:

    def __init__(self, location):

        pages = 1
    
        while True:
            param={
                "apiKey" : NEWS_API_KEY,
                "q" : location,
                "language" : "en",
                "pageSize" : 5,
                "page" : pages
            }

            response = requests.get(NEWS_BASELINE+NEW_EVERYTHING_END_POINT, params=param)
            data = response.json()
            option_input = int(input(
                    "Choose a headline in the menu...\n"
                    "1." + data["articles"][0]["title"] + "\n"
                    "2." + data["articles"][1]["title"] + "\n"
                    "3." + data["articles"][2]["title"] + "\n"
                    "4." + data["articles"][3]["title"] + "\n"
                    "5." + data["articles"][4]["title"] + "\n"
                    "6. Next " + "\n"
                    "Option : "))
            if( 6 > option_input > 0):
                self.url_string = data["articles"][option_input-1]["url"]
                self.headline = data["articles"][option_input-1]["title"]
                self.format_news = [
                    ("Headline :" , self.headline),
                    ("URL :" , self.url_string)
                ]
                print(self.format_news)
                break
            elif option_input == 6:
                pages += 1
                continue

    def news_data(self) -> list:
        return self.format_news
                  
#NOTE: Save history record
class History:
    
    def saved_history(self, input_option, rows):
        
        path = os.path.join(os.path.dirname(__file__), "history_record.xlsx")

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active

        if input_option == 2 or input_option == 3:
            ws.title = "History"

        ws.insert_rows(1, amount=len(rows) + 1) 

        for i, (label, value) in enumerate(rows, start=1):  
            ws.cell(row=i, column=1, value=label) 
            ws.cell(row=i, column=2, value=value)  
        
        wb.save(path)
        self.path = path

    def read_history(self):
        file_path = os.path.join(os.path.dirname(__file__), "history_record.xlsx")

        if not os.path.exists(file_path):
            print("History file doesn't exist!!")
            return
        wb = openpyxl.load_workbook(file_path)
        sh = wb.active

        rows = list(sh.iter_rows(values_only=True)) 

        if not rows or all(all(cell is None for cell in row)for row in rows):
            print("No history data found!!")
            wb.close()
            return
        
        print("\n--History--\n")
        for row in rows:
            option = row[0] if len(row) > 0 and row[0] is not None else " "
            data = row[1] if len(row) > 1 and row[1] is not None else ""
            print(f"{option} {data}")
    
        wb.close()

if __name__ == "__main__":
   Dashboard()
  